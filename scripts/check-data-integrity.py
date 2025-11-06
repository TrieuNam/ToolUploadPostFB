#!/usr/bin/env python3
"""
Kiểm tra data integrity giữa JSON và Excel
"""

import json
import sys
from openpyxl import load_workbook

def check_data_integrity(json_file, excel_file):
    """
    So sánh data giữa JSON và Excel
    """
    print("\n" + "="*60)
    print("🔍 KIỂM TRA DATA INTEGRITY")
    print("="*60)
    
    # Load JSON
    with open(json_file, 'r', encoding='utf-8') as f:
        json_data = json.load(f)
    
    json_videos = json_data.get('videos', [])
    json_urls = [v.get('url', '') for v in json_videos]
    
    print(f"\n📄 JSON file:")
    print(f"   Total videos: {len(json_videos)}")
    print(f"   File: {json_file}")
    
    # Load Excel
    wb = load_workbook(excel_file)
    ws = wb.active
    
    excel_urls = []
    excel_new_urls = []
    
    for row in ws.iter_rows(min_row=2, max_col=17, values_only=True):
        if row[1]:  # video_download_url
            excel_urls.append(row[1])
            if row[7] == 'NEW':  # status
                excel_new_urls.append(row[1])
    
    print(f"\n📊 Excel file:")
    print(f"   Total videos: {len(excel_urls)}")
    print(f"   Videos NEW: {len(excel_new_urls)}")
    print(f"   File: {excel_file}")
    
    # So sánh
    print(f"\n" + "="*60)
    print("📋 KẾT QUẢ SO SÁNH")
    print("="*60)
    
    # Videos trong JSON nhưng không có trong Excel
    missing_in_excel = []
    for url in json_urls:
        if url not in excel_urls:
            missing_in_excel.append(url)
    
    # Videos trong Excel nhưng không có trong JSON (từ JSON này)
    extra_in_excel = []
    for url in excel_new_urls:
        if url not in json_urls:
            extra_in_excel.append(url)
    
    if len(missing_in_excel) == 0 and len(extra_in_excel) == 0:
        print("\n✅ PERFECT! Data khớp 100%")
        print(f"   {len(json_videos)} videos trong JSON")
        print(f"   {len(excel_new_urls)} videos NEW trong Excel")
        print(f"   ✅ Không có video nào bị miss!")
    else:
        if missing_in_excel:
            print(f"\n⚠️ MISSING IN EXCEL: {len(missing_in_excel)} videos")
            for idx, url in enumerate(missing_in_excel[:5], 1):
                print(f"   {idx}. {url[:80]}...")
            if len(missing_in_excel) > 5:
                print(f"   ... và {len(missing_in_excel) - 5} videos khác")
        
        if extra_in_excel:
            print(f"\n📌 EXTRA IN EXCEL: {len(extra_in_excel)} videos")
            print("   (Có thể từ lần import trước)")
    
    # Chi tiết thống kê
    print(f"\n" + "="*60)
    print("📊 THỐNG KÊ CHI TIẾT")
    print("="*60)
    print(f"JSON:")
    print(f"   Total: {len(json_videos)} videos")
    print(f"\nExcel:")
    print(f"   Total all: {len(excel_urls)} videos")
    print(f"   Status NEW: {len(excel_new_urls)} videos")
    print(f"   Status POSTED: {len(excel_urls) - len(excel_new_urls)} videos")
    
    # Sample data
    print(f"\n" + "="*60)
    print("📝 SAMPLE DATA (5 videos đầu tiên)")
    print("="*60)
    
    count = 0
    for idx, row in enumerate(ws.iter_rows(min_row=2, max_col=17, values_only=True), 2):
        if row[7] == 'NEW' and count < 5:
            print(f"\nRow {idx}:")
            print(f"   ID: {row[0]}")
            print(f"   Title: {row[2]}")
            print(f"   URL: {row[1][:60]}...")
            print(f"   Status: {row[7]}")
            print(f"   Hashtags: {row[4]}")
            print(f"   Shopee: {row[5][:50] if row[5] else ''}...")
            count += 1
    
    print("\n" + "="*60)
    
    return {
        'json_count': len(json_videos),
        'excel_total': len(excel_urls),
        'excel_new': len(excel_new_urls),
        'missing': len(missing_in_excel),
        'extra': len(extra_in_excel)
    }


if __name__ == "__main__":
    json_file = "data/tiktok-videos-1762452044856.json"
    excel_file = "data/posts.xlsx"
    
    if len(sys.argv) > 1:
        json_file = sys.argv[1]
    
    result = check_data_integrity(json_file, excel_file)
    
    print("\n✨ Check hoàn tất!")
