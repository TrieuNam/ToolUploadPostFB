from openpyxl import load_workbook

wb = load_workbook('data/posts.xlsx')
ws = wb.active

print(f"Total rows: {ws.max_row}")

new_count = 0
for row in ws.iter_rows(min_row=2, max_col=8, values_only=True):
    if row[7] == 'NEW':
        new_count += 1

print(f"Videos với status=NEW: {new_count}")
print("\nSample NEW videos (first 5):")

count = 0
for idx, row in enumerate(ws.iter_rows(min_row=2, max_col=17, values_only=True), 2):
    if row[7] == 'NEW' and count < 5:
        print(f"  Row {idx}: ID={row[0]}, Title={row[2]}, Status={row[7]}")
        count += 1
