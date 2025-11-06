#!/usr/bin/env python3
"""
TikTok Video Scraper
Lấy danh sách link video từ TikTok profile và thêm vào Excel
"""

import requests
import json
import time
from datetime import datetime
import openpyxl
from openpyxl import load_workbook
import os

class TikTokScraper:
    def __init__(self, excel_path):
        self.excel_path = excel_path
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'vi-VN,vi;q=0.9,en-US;q=0.8,en;q=0.7',
            'Referer': 'https://www.tiktok.com/'
        }
    
    def get_user_videos(self, username, max_videos=10):
        """
        Lấy danh sách video từ TikTok profile
        
        Args:
            username: TikTok username (vd: @cartonvn)
            max_videos: Số lượng video tối đa cần lấy
        
        Returns:
            List of video URLs
        """
        print(f"\n🔍 Đang quét TikTok profile: @{username}")
        print(f"📊 Số video tối đa: {max_videos}")
        
        # Remove @ if exists
        username = username.replace('@', '')
        
        # TikTok profile URL
        profile_url = f"https://www.tiktok.com/@{username}"
        
        try:
            # Method 1: Sử dụng TikTok web scraping
            videos = self._scrape_from_web(username, max_videos)
            
            if videos:
                print(f"✅ Tìm thấy {len(videos)} video!")
                return videos
            else:
                print("⚠️ Không tìm thấy video. Thử method khác...")
                # Method 2: Manual input (fallback)
                return self._manual_input()
                
        except Exception as e:
            print(f"❌ Lỗi: {e}")
            return self._manual_input()
    
    def _scrape_from_web(self, username, max_videos):
        """
        Scrape videos từ TikTok web (simplified version)
        Lưu ý: TikTok có thể block scraping, nên cần dùng API hoặc tool khác
        """
        videos = []
        
        # Placeholder - In thực tế cần dùng:
        # - Playwright/Selenium để render JavaScript
        # - TikTok API (cần đăng ký)
        # - Third-party API như RapidAPI
        
        print("⚠️ TikTok web scraping phức tạp do JavaScript rendering.")
        print("📝 Vui lòng nhập link video thủ công hoặc dùng extension.")
        
        return videos
    
    def _manual_input(self):
        """
        Nhập link video thủ công
        """
        print("\n" + "="*50)
        print("📝 NHẬP LINK VIDEO THỦ CÔNG")
        print("="*50)
        print("Hướng dẫn:")
        print("1. Mở TikTok profile trong browser")
        print("2. Copy link từng video (right click → Copy link)")
        print("3. Paste vào đây (mỗi link 1 dòng)")
        print("4. Gõ 'done' khi xong")
        print("="*50 + "\n")
        
        videos = []
        count = 1
        
        while True:
            link = input(f"Video {count} (hoặc 'done'): ").strip()
            
            if link.lower() == 'done':
                break
            
            if link and 'tiktok.com' in link:
                videos.append(link)
                print(f"✅ Đã thêm video {count}")
                count += 1
            elif link:
                print("⚠️ Link không hợp lệ. Vui lòng nhập link TikTok!")
        
        return videos
    
    def extract_video_id(self, url):
        """
        Trích xuất video ID từ TikTok URL
        URL format: https://www.tiktok.com/@username/video/7569580062823922583
        """
        try:
            if '/video/' in url:
                video_id = url.split('/video/')[1].split('?')[0].split('/')[0]
                return video_id
            return None
        except:
            return None
    
    def add_videos_to_excel(self, video_urls, title_template="Video từ TikTok", description="", hashtags="", shopee_links=""):
        """
        Thêm danh sách video vào Excel
        
        Args:
            video_urls: List các URL video
            title_template: Template cho title (sẽ thêm số thứ tự)
            description: Mô tả chung
            hashtags: Hashtags chung
            shopee_links: Shopee links chung
        """
        print(f"\n📝 Đang thêm {len(video_urls)} video vào Excel...")
        
        try:
            # Load Excel
            if os.path.exists(self.excel_path):
                wb = load_workbook(self.excel_path)
                ws = wb.active
            else:
                # Tạo file mới nếu chưa có
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Posts"
                
                # Tạo headers
                headers = [
                    'id', 'video_download_url', 'title', 'description', 'hashtags',
                    'shopee_links', 'scheduled_time', 'status', 'local_video_url',
                    'local_video_path', 'video_size', 'tiktok_url', 'tiktok_post_id',
                    'facebook_post_id', 'facebook_post_url', 'facebook_posted_at', 'error_message'
                ]
                ws.append(headers)
            
            # Lấy ID cuối cùng
            max_id = 0
            for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                if row[0] and isinstance(row[0], int):
                    max_id = max(max_id, row[0])
            
            # Thêm từng video
            added_count = 0
            for idx, url in enumerate(video_urls, 1):
                video_id = self.extract_video_id(url)
                
                # Check duplicate
                is_duplicate = False
                for row in ws.iter_rows(min_row=2, max_col=12, values_only=True):
                    if row[1] == url or row[11] == url:  # video_download_url or tiktok_url
                        is_duplicate = True
                        break
                
                if is_duplicate:
                    print(f"⚠️ Video {idx} đã tồn tại, bỏ qua...")
                    continue
                
                # Tạo row mới
                new_row = [
                    max_id + idx,                           # id
                    url,                                     # video_download_url
                    f"{title_template} #{max_id + idx}",    # title
                    description,                             # description
                    hashtags,                                # hashtags
                    shopee_links,                            # shopee_links
                    "",                                      # scheduled_time (empty = post ngay)
                    "NEW",                                   # status
                    "",                                      # local_video_url
                    "",                                      # local_video_path
                    "",                                      # video_size
                    url,                                     # tiktok_url
                    video_id or "",                          # tiktok_post_id
                    "",                                      # facebook_post_id
                    "",                                      # facebook_post_url
                    "",                                      # facebook_posted_at
                    ""                                       # error_message
                ]
                
                ws.append(new_row)
                added_count += 1
                print(f"✅ Đã thêm: {new_row[2]} ({url[:50]}...)")
            
            # Save Excel
            wb.save(self.excel_path)
            print(f"\n🎉 Hoàn thành! Đã thêm {added_count} video vào Excel")
            print(f"📁 File: {self.excel_path}")
            
            return added_count
            
        except Exception as e:
            print(f"❌ Lỗi khi thêm vào Excel: {e}")
            return 0
    
    def interactive_mode(self):
        """
        Chế độ tương tác
        """
        print("\n" + "="*60)
        print("🎬 TIKTOK VIDEO SCRAPER - INTERACTIVE MODE")
        print("="*60)
        
        # Nhập username
        username = input("\n👤 Nhập TikTok username (vd: cartonvn): ").strip()
        if not username:
            print("❌ Username không được để trống!")
            return
        
        # Nhập số video
        try:
            max_videos = int(input("📊 Số video cần lấy (mặc định 10): ").strip() or "10")
        except:
            max_videos = 10
        
        # Lấy video URLs
        video_urls = self.get_user_videos(username, max_videos)
        
        if not video_urls:
            print("❌ Không có video nào được thêm.")
            return
        
        # Nhập thông tin chung
        print("\n" + "="*60)
        print("📝 THÔNG TIN CHUNG CHO TẤT CẢ VIDEO")
        print("="*60)
        
        title_template = input("📌 Title template (mặc định: 'Video từ TikTok'): ").strip() or "Video từ TikTok"
        description = input("📄 Description chung: ").strip()
        hashtags = input("🏷️  Hashtags (vd: #viral #trending): ").strip()
        shopee_links = input("🛒 Shopee links (cách nhau bởi dấu phẩy): ").strip()
        
        # Thêm vào Excel
        self.add_videos_to_excel(
            video_urls,
            title_template=title_template,
            description=description,
            hashtags=hashtags,
            shopee_links=shopee_links
        )


def main():
    """
    Main function
    """
    # Path to Excel
    excel_path = os.path.join(
        os.path.dirname(__file__),
        '..',
        'data',
        'posts.xlsx'
    )
    
    # Create scraper
    scraper = TikTokScraper(excel_path)
    
    # Run interactive mode
    scraper.interactive_mode()
    
    print("\n✨ Xong! Bây giờ bạn có thể chạy workflow để download và post.")
    print("📖 Hướng dẫn: docs/USAGE.md")


if __name__ == "__main__":
    main()
