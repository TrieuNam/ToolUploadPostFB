#!/usr/bin/env python3
"""
Fix Excel structure - Thêm cột description
"""

from openpyxl import load_workbook
import sys

def fix_excel_structure(excel_file):
    """
    Thêm cột description vào Excel (sau title, trước hashtags)
    """
    print("\n" + "="*60)
    print("🔧 FIX EXCEL STRUCTURE - THÊM CỘT DESCRIPTION")
    print("="*60)
    
    try:
        # Load Excel
        wb = load_workbook(excel_file)
        ws = wb.active
        
        # Current structure
        current_headers = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]
        
        print(f"\n📊 Cấu trúc hiện tại:")
        print(f"   Columns: {len(current_headers)}")
        for idx, h in enumerate(current_headers, 1):
            print(f"   {idx}. {h}")
        
        # Check if description already exists
        if 'description' in current_headers:
            print("\n✅ Cột 'description' đã tồn tại!")
            return
        
        print(f"\n🔧 Thêm cột 'description' vào vị trí 4 (sau 'title')...")
        
        # Insert column after 'title' (position 4)
        ws.insert_cols(4)
        
        # Set header
        ws.cell(1, 4).value = 'description'
        
        # New structure
        new_headers = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]
        
        print(f"\n✅ Cấu trúc mới:")
        print(f"   Columns: {len(new_headers)}")
        for idx, h in enumerate(new_headers, 1):
            marker = " ← NEW" if h == 'description' else ""
            print(f"   {idx}. {h}{marker}")
        
        # Save
        backup_file = excel_file.replace('.xlsx', '_backup.xlsx')
        wb.save(backup_file)
        print(f"\n💾 Backup: {backup_file}")
        
        wb.save(excel_file)
        print(f"💾 Saved: {excel_file}")
        
        print(f"\n🎉 HOÀN THÀNH!")
        print(f"   ✅ Đã thêm cột 'description'")
        print(f"   ✅ Tất cả data giữ nguyên")
        print(f"   ✅ Backup đã được tạo")
        
        return True
        
    except Exception as e:
        print(f"\n❌ Lỗi: {e}")
        return False


if __name__ == "__main__":
    excel_file = "data/posts.xlsx"
    
    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
    
    fix_excel_structure(excel_file)
