#!/usr/bin/env python3
"""
Import TikTok videos từ JSON vào Excel
"""

import json
import sys
import os
from openpyxl import load_workbook
from datetime import datetime

def import_json_to_excel(json_file, excel_file):
    """
    Import videos từ JSON file vào Excel
    
    Args:
        json_file: Path to JSON file từ extension
        excel_file: Path to Excel file
    """
    print(f"\n📥 Đang import JSON vào Excel...")
    print(f"📄 JSON: {json_file}")
    print(f"📊 Excel: {excel_file}")
    
    try:
        # Load JSON
        with open(json_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        videos = data.get('videos', [])
        total = data.get('total', len(videos))
        
        print(f"\n✅ Đọc được {len(videos)} videos từ JSON")
        
        if len(videos) == 0:
            print("❌ Không có video nào trong JSON!")
            return 0
        
        # Load Excel
        if os.path.exists(excel_file):
            wb = load_workbook(excel_file)
            ws = wb.active
            print(f"✅ Đã mở Excel file")
        else:
            print(f"❌ Excel file không tồn tại: {excel_file}")
            return 0
        
        # Lấy ID cuối cùng
        max_id = 0
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0] and isinstance(row[0], (int, float)):
                max_id = max(max_id, int(row[0]))
        
        print(f"📝 ID hiện tại: {max_id}")
        
        # Nhập thông tin chung
        print("\n" + "="*60)
        print("📝 THÔNG TIN CHUNG CHO TẤT CẢ VIDEO")
        print("="*60)
        
        title_template = input("📌 Title template (vd: 'Thor phim hay'): ").strip() or "Video từ TikTok"
        description = input("📄 Description chung: ").strip()
        hashtags = input("🏷️  Hashtags (vd: #viral #trending): ").strip()
        shopee_links = input("🛒 Shopee links (cách nhau bởi dấu phẩy): ").strip()
        
        # Thêm từng video
        added_count = 0
        skipped_count = 0
        
        print("\n" + "="*60)
        print("📥 ĐANG THÊM VIDEO VÀO EXCEL...")
        print("="*60)
        
        for idx, video in enumerate(videos, 1):
            url = video.get('url', '')
            video_id = video.get('video_id', '')
            
            if not url:
                print(f"⚠️ Video {idx}: Không có URL, bỏ qua...")
                skipped_count += 1
                continue
            
            # Check duplicate
            is_duplicate = False
            for row in ws.iter_rows(min_row=2, max_col=16, values_only=True):
                if row[1] == url or row[10] == url:  # video_download_url (col 2) or tiktok_url (col 11)
                    is_duplicate = True
                    break
            
            if is_duplicate:
                print(f"⚠️ Video {idx}: Đã tồn tại, bỏ qua... ({url[:50]}...)")
                skipped_count += 1
                continue
            
            # Tạo row mới - Match với Excel structure CÓ description
            # Excel headers: id, video_download_url, title, description, hashtags, 
            #                shopee_links, status, local_video_url, local_video_path, 
            #                video_size, tiktok_url, error_message, facebook_post_id, 
            #                facebook_post_url, facebook_posted_at, scheduled_time
            new_id = max_id + added_count + 1
            new_row = [
                new_id,                              # 1. id
                url,                                 # 2. video_download_url
                f"{title_template} #{new_id}",      # 3. title
                description,                         # 4. description ← NEW
                hashtags,                            # 5. hashtags
                shopee_links,                        # 6. shopee_links
                "NEW",                               # 7. status
                "",                                  # 8. local_video_url
                "",                                  # 9. local_video_path
                "",                                  # 10. video_size
                url,                                 # 11. tiktok_url
                "",                                  # 12. error_message
                "",                                  # 13. facebook_post_id
                "",                                  # 14. facebook_post_url
                "",                                  # 15. facebook_posted_at
                ""                                   # 16. scheduled_time (empty = post ngay)
            ]
            
            ws.append(new_row)
            added_count += 1
            print(f"✅ Video {idx}/{len(videos)}: {new_row[2]}")
        
        # Save Excel
        wb.save(excel_file)
        
        print("\n" + "="*60)
        print("🎉 HOÀN THÀNH!")
        print("="*60)
        print(f"✅ Đã thêm: {added_count} videos")
        print(f"⚠️ Đã bỏ qua: {skipped_count} videos (duplicate)")
        print(f"📁 Excel: {excel_file}")
        print("\n🚀 Bây giờ bạn có thể chạy automation workflow!")
        
        return added_count
        
    except FileNotFoundError:
        print(f"❌ Lỗi: Không tìm thấy file {json_file}")
        return 0
    except json.JSONDecodeError:
        print(f"❌ Lỗi: File JSON không hợp lệ")
        return 0
    except Exception as e:
        print(f"❌ Lỗi: {e}")
        return 0


def main():
    """
    Main function
    """
    print("\n" + "="*60)
    print("🎬 IMPORT TIKTOK VIDEOS FROM JSON TO EXCEL")
    print("="*60)
    
    # Get paths
    if len(sys.argv) > 1:
        # JSON file từ command line
        json_file = sys.argv[1]
    else:
        # Hỏi user
        json_file = input("\n📄 Nhập path JSON file (hoặc kéo thả file vào đây): ").strip().strip('"')
    
    # Default Excel path
    excel_file = os.path.join(
        os.path.dirname(__file__),
        '..',
        'data',
        'posts.xlsx'
    )
    
    # Check files exist
    if not os.path.exists(json_file):
        print(f"\n❌ File không tồn tại: {json_file}")
        print("\n💡 Hướng dẫn:")
        print("   1. Mở TikTok profile trong browser")
        print("   2. Click extension 'Extract Videos'")
        print("   3. Click 'Download JSON'")
        print("   4. Chạy lại script này với path đến file JSON")
        return
    
    if not os.path.exists(excel_file):
        print(f"\n❌ Excel file không tồn tại: {excel_file}")
        return
    
    # Import
    result = import_json_to_excel(json_file, excel_file)
    
    if result > 0:
        print("\n✨ Success!")
    else:
        print("\n❌ Không thêm được video nào.")


if __name__ == "__main__":
    main()
